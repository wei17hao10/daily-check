__author__ = 'Wayne'
import os
from openpyxl import Workbook


def get_filename(dirname):
	wb = Workbook()
	ws = wb.create_sheet(title="Data")
	ws["A1"] = "File Path"
	ws["B1"] = "Hyperlink"
	for roots, dirs, files in os.walk(dirname):
		dirs.sort()
		files.sort()
		rowA = roots.count('/')*'|  '+'|=>'+roots.split('/')[-1]
		ws.append([rowA])
		for file in files:
			if file[0] != '.':
				rowA = (roots.count('/')+1)*'|  '+'|--'+file
				rowB = '=hyperlink("' + roots + '/' + file + '","' + file + '")'
				ws.append([rowA, rowB])
	del wb['Sheet']
	wb.save("FileTrace.xlsx") 


print('script is running now\nPlease check the result file --> FileTrace.xlsx')
get_filename('./')

